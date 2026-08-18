import openpyxl
import pytest
from pydantic import ValidationError

from app.config import Settings
from app.core.analyzer import IncrementalAnalyzer
from app.core.extractor import extract_file_text


def test_tabular_config_validation():
    """Verify numeric bounds for TABULAR_MAX_SHEETS, TABULAR_MAX_ROWS, and TABULAR_MAX_CHARACTERS."""
    # Defaults
    settings = Settings()
    assert settings.TABULAR_MAX_SHEETS == 10
    assert settings.TABULAR_MAX_ROWS == 10000
    assert settings.TABULAR_MAX_CHARACTERS == 50000
    assert settings.MAX_SPREADSHEET_SHEETS == 10
    assert settings.MAX_SPREADSHEET_ROWS == 10000
    assert settings.MAX_SPREADSHEET_CHARACTERS == 50000

    # Custom valid values
    custom = Settings(
        TABULAR_MAX_SHEETS=5,
        TABULAR_MAX_ROWS=100,
        TABULAR_MAX_CHARACTERS=2000,
    )
    assert custom.TABULAR_MAX_SHEETS == 5
    assert custom.TABULAR_MAX_ROWS == 100
    assert custom.TABULAR_MAX_CHARACTERS == 2000

    # Invalid non-positive values
    with pytest.raises(ValidationError):
        Settings(TABULAR_MAX_SHEETS=0)
    with pytest.raises(ValidationError):
        Settings(TABULAR_MAX_ROWS=-10)
    with pytest.raises(ValidationError):
        Settings(TABULAR_MAX_CHARACTERS=0)


def test_csv_streaming_row_and_char_limits(tmp_path):
    """Verify CSV extraction streams row-by-row and enforces max row and character bounds."""
    csv_file = tmp_path / "test_data.csv"
    
    # Write a CSV with 50 rows
    lines = ["header_col1,header_col2"]
    for i in range(1, 50):
        lines.append(f"row{i}_val1,row{i}_val2")
    csv_file.write_text("\n".join(lines), encoding="utf-8")

    # 1. Test row limit: max 5 rows (header + 4 data rows)
    settings_row_limit = Settings(TABULAR_MAX_ROWS=5, TABULAR_MAX_CHARACTERS=10000)
    text = extract_file_text(str(csv_file), settings=settings_row_limit)
    
    assert "header_col1 header_col2" in text
    assert "row1_val1 row1_val2" in text
    assert "row4_val1 row4_val2" in text
    assert "row5_val1" not in text
    assert "row10_val1" not in text

    # 2. Test character limit
    settings_char_limit = Settings(TABULAR_MAX_ROWS=100, TABULAR_MAX_CHARACTERS=35)
    text_char = extract_file_text(str(csv_file), settings=settings_char_limit)
    
    assert len(text_char) <= 35
    assert text_char.startswith("header_col1 header_col2")


def test_excel_streaming_sheet_row_and_char_limits(tmp_path):
    """Verify Excel extraction uses read-only mode and respects sheet, row, and character bounds."""
    xlsx_file = tmp_path / "test_workbook.xlsx"

    # Create a workbook with 4 sheets and 30 rows per sheet
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    default_sheet.title = "Sheet1"

    sheet_names = ["Sheet1", "Sheet2", "Sheet3", "Sheet4"]
    for sname in sheet_names:
        if sname == "Sheet1":
            ws = default_sheet
        else:
            ws = wb.create_sheet(title=sname)
        
        ws.append(["ColA", "ColB", "ColC"])
        for r in range(1, 30):
            ws.append([f"{sname}_R{r}_A", f"{sname}_R{r}_B", f"{sname}_R{r}_C"])

    wb.save(xlsx_file)
    wb.close()

    # 1. Test sheet and row limits: max 2 sheets, max 3 rows per sheet
    settings_limits = Settings(
        TABULAR_MAX_SHEETS=2,
        TABULAR_MAX_ROWS=3,
        TABULAR_MAX_CHARACTERS=50000,
    )
    extracted_text = extract_file_text(str(xlsx_file), settings=settings_limits)

    # Should contain Sheet1 and Sheet2, but not Sheet3 or Sheet4
    assert "Sheet1_R1_A" in extracted_text
    assert "Sheet1_R2_A" in extracted_text
    assert "Sheet1_R3_A" not in extracted_text  # Row 1 is header, Row 2 is R1, Row 3 is R2 -> total 3 rows
    assert "Sheet2_R1_A" in extracted_text
    assert "Sheet3_R1_A" not in extracted_text
    assert "Sheet4_R1_A" not in extracted_text

    # 2. Test character limit
    settings_char = Settings(
        TABULAR_MAX_SHEETS=10,
        TABULAR_MAX_ROWS=100,
        TABULAR_MAX_CHARACTERS=50,
    )
    extracted_char = extract_file_text(str(xlsx_file), settings=settings_char)

    assert len(extracted_char) <= 50
    assert extracted_char.startswith("ColA ColB ColC")


def test_downstream_topic_analysis_with_truncated_tabular_text(tmp_path):
    """Verify that IncrementalAnalyzer processes truncated tabular extraction outputs without error."""
    # Create CSV and Excel files
    csv_file = tmp_path / "finance_data.csv"
    csv_lines = ["date,amount,category,description"]
    for i in range(100):
        csv_lines.append(f"2026-01-{i%28+1:02d},{100+i},finance,quarterly revenue report entry {i}")
    csv_file.write_text("\n".join(csv_lines), encoding="utf-8")

    xlsx_file = tmp_path / "clinical_trials.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trials"
    ws.append(["study_id", "patient_count", "phase", "indication"])
    for i in range(100):
        ws.append([f"STUDY-{i}", f"{10+i}", "Phase III", "Oncology clinical trial protocol"])
    wb.save(xlsx_file)
    wb.close()

    settings = Settings(
        TABULAR_MAX_SHEETS=1,
        TABULAR_MAX_ROWS=10,
        TABULAR_MAX_CHARACTERS=200,
    )

    csv_text = extract_file_text(str(csv_file), settings=settings)
    xlsx_text = extract_file_text(str(xlsx_file), settings=settings)

    # Ensure output is truncated and non-empty
    assert len(csv_text) <= 200
    assert len(xlsx_text) <= 200
    assert "finance" in csv_text or "date" in csv_text
    assert "study_id" in xlsx_text or "Oncology" in xlsx_text

    # Feed into IncrementalAnalyzer
    from unittest.mock import MagicMock
    mock_db = MagicMock()
    mock_db.get_all_embeddings.return_value = {}

    corpus = {
        "finance_data.csv": csv_text,
        "clinical_trials.xlsx": xlsx_text,
    }

    analyzer = IncrementalAnalyzer(
        max_folders=2, stop_words={"the", "and"}, db=mock_db, model_path="all-MiniLM-L6-v2"
    )
    analyzer.partial_fit(str(tmp_path), corpus)
    plan = analyzer.generate_sorting_plan(str(tmp_path))

    # Plan should successfully partition without raising any exception
    assert isinstance(plan, dict)
    assert len(analyzer.corpus) == 2
